from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
import argparse
import sys
import re

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None


def parse_args():
    parser = argparse.ArgumentParser(
        description="Tổng hợp report từ file xlsx và pdf trong cùng thư mục."
    )
    parser.add_argument(
        "--folder",
        required=True,
        type=str,
        help="Đường dẫn thư mục chứa file báo cáo",
    )
    parser.add_argument(
        "--top-k",
        required=True,
        type=int,
        help="Số lượng top muốn hiển thị cho phần events (.xlsx)",
    )

    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists() or not folder.is_dir():
        parser.error(f"Thư mục không tồn tại hoặc không hợp lệ: {folder}")

    if args.top_k <= 0:
        parser.error("--top-k phải là số nguyên dương.")

    return folder, args.top_k


def build_output_file(folder_path: Path) -> Path:
    folder_name = folder_path.name.strip()
    if not folder_name:
        folder_name = "default"
    return folder_path / f"{folder_name}_report.txt"


def normalize_cell_value(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def parse_int_with_dots(text: str) -> int:
    text = text.strip().replace(".", "").replace(",", "")
    return int(text) if text.isdigit() else 0


def process_excel_file(file_path: Path):
    counter_domain = Counter()
    counter_ip = Counter()
    counter_url = Counter()
    ip_nation_map = defaultdict(Counter)
    total_rows = 0

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=12, min_col=4, max_col=8, values_only=True):
                val_domain = normalize_cell_value(row[0])  # D
                val_ip = normalize_cell_value(row[1])      # E
                val_nation = normalize_cell_value(row[2])  # F
                val_url = normalize_cell_value(row[4])     # H

                if (
                    val_domain is not None
                    or val_ip is not None
                    or val_nation is not None
                    or val_url is not None
                ):
                    total_rows += 1

                if val_domain is not None:
                    counter_domain.update([val_domain])

                if val_ip is not None:
                    counter_ip.update([val_ip])

                if val_url is not None:
                    counter_url.update([val_url])

                if val_ip is not None and val_nation is not None:
                    ip_nation_map[val_ip].update([val_nation])
    finally:
        wb.close()

    return counter_domain, counter_ip, counter_url, ip_nation_map, total_rows


def extract_pdf_text(file_path: Path) -> str:
    if PdfReader is None:
        raise RuntimeError("Thiếu thư viện pypdf. Cài bằng: pip install pypdf")

    reader = PdfReader(str(file_path))
    texts = []
    for page in reader.pages:
        texts.append(page.extract_text() or "")
    return "\n".join(texts)


def normalize_pdf_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\r", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def extract_section(text: str, start_marker: str, end_markers: list[str]) -> str:
    start_idx = text.find(start_marker)
    if start_idx == -1:
        return ""

    section = text[start_idx + len(start_marker):]

    end_positions = []
    for marker in end_markers:
        idx = section.find(marker)
        if idx != -1:
            end_positions.append(idx)

    if end_positions:
        section = section[:min(end_positions)]

    return section.strip()


def parse_attack_lines(block_text: str):
    results = []
    pattern = re.compile(r"([A-Za-z0-9_\-./]+)\s+(\d+(?:,\d+)?)%\s+\(([\d.]+)\)")

    for line in block_text.splitlines():
        line = line.strip()
        if not line:
            continue

        m = pattern.search(line)
        if not m:
            continue

        attack_type = m.group(1).strip()
        percent_str = m.group(2).replace(",", ".")
        count_str = m.group(3).strip()

        try:
            percent = float(percent_str)
        except ValueError:
            percent = 0.0

        count = parse_int_with_dots(count_str)

        results.append({
            "type": attack_type,
            "percent": percent,
            "count": count,
        })

    return results


def process_pdf_file(file_path: Path):
    raw_text = extract_pdf_text(file_path)
    text = normalize_pdf_text(raw_text)

    web_total = 0
    ddos_total = 0

    m_web = re.search(
        r"Tấn công lỗ hổng web\s*\(Tổng số\s*([\d.]+)\s*cuộc tấn công\)",
        text,
        flags=re.IGNORECASE,
    )
    if m_web:
        web_total = parse_int_with_dots(m_web.group(1))

    if re.search(
        r"Không có dữ liệu tấn công DDOS tầng ứng dụng để hiển thị",
        text,
        flags=re.IGNORECASE,
    ):
        ddos_total = 0
    else:
        m_ddos = re.search(
            r"Tấn công DDOS tầng ứng dụng\s*\(Tổng số\s*([\d.]+)\s*cuộc tấn công\)",
            text,
            flags=re.IGNORECASE,
        )
        if m_ddos:
            ddos_total = parse_int_with_dots(m_ddos.group(1))

    web_attack_block = extract_section(
        text,
        start_marker="Các loại tấn công khai thác lỗ hổng",
        end_markers=[
            "DDOS TẦNG ỨNG DỤNG",
            "Các loại tấn công DDOS",
            "TƯỜNG LỬA BẢO VỆ WEBSITE Tìm hiểu thêm",
            "Top tên miền bị tấn công khai thác lỗ hổng",
        ],
    )
    web_attack_types = parse_attack_lines(web_attack_block)

    ddos_attack_block = extract_section(
        text,
        start_marker="Các loại tấn công DDOS",
        end_markers=[
            "TƯỜNG LỬA BẢO VỆ WEBSITE Tìm hiểu thêm",
            "Top tên miền bị tấn công DDOS L7",
            "Mọi chi tiết xin liên hệ",
        ],
    )

    if re.search(r"Không có dữ liệu để hiển thị", ddos_attack_block, flags=re.IGNORECASE):
        ddos_attack_types = []
    else:
        ddos_attack_types = parse_attack_lines(ddos_attack_block)

    if web_total == 0 and web_attack_types:
        web_total = sum(item["count"] for item in web_attack_types)

    if ddos_total == 0 and ddos_attack_types:
        ddos_total = sum(item["count"] for item in ddos_attack_types)

    return {
        "web_total": web_total,
        "ddos_total": ddos_total,
        "web_attack_types": web_attack_types,
        "ddos_attack_types": ddos_attack_types,
    }


def get_top_ip_with_nation(counter_ip: Counter, ip_nation_map: dict, top_k: int, total_events: int):
    result = []
    for ip, count in counter_ip.most_common(top_k):
        nation_counter = ip_nation_map.get(ip, Counter())
        if nation_counter:
            nation, _ = nation_counter.most_common(1)[0]
        else:
            nation = "Unknown"

        percent = (count / total_events * 100) if total_events > 0 else 0.0

        result.append({
            "ip": ip,
            "count": count,
            "nation": nation,
            "percent": percent,
        })
    return result


def format_top_ip_with_nation(top_ip_data):
    lines = []
    title = "TOP IPs và quốc gia tương ứng"
    lines.append(title)
    lines.append("-" * len(title))

    if not top_ip_data:
        lines.append("Không có dữ liệu.")
        return lines

    for idx, item in enumerate(top_ip_data, start=1):
        percent_str = f"{item['percent']:.2f}".replace(".", ",")
        lines.append(
            f"{idx:>2}. IP: {item['ip']} -> {item['count']} ({percent_str}%) | Nation: {item['nation']}"
        )
    return lines


def format_top(counter: Counter, title: str, top_k: int):
    lines = []
    lines.append(title)
    lines.append("-" * len(title))

    if not counter:
        lines.append("Không có dữ liệu.")
        return lines

    for idx, (value, count) in enumerate(counter.most_common(top_k), start=1):
        lines.append(f"{idx:>2}. {value} -> {count}")
    return lines


def format_all_attack_types(title: str, attack_counter: Counter):
    lines = []
    lines.append(title)
    lines.append("-" * len(title))

    if not attack_counter:
        lines.append("Không có dữ liệu.")
        return lines

    total_attack_type_count = sum(attack_counter.values())

    for idx, (attack_type, count) in enumerate(attack_counter.most_common(), start=1):
        percent = (count / total_attack_type_count * 100) if total_attack_type_count > 0 else 0.0
        percent_str = f"{percent:.2f}".replace(".", ",")
        lines.append(f"{idx:>2}. {attack_type} -> {percent_str}% ({count})")

    return lines


def handle_events(folder_path: Path, top_k: int):
    excel_files = list(folder_path.rglob("*.xlsx"))

    section_lines = []
    section_lines.append("=" * 60)
    section_lines.append("EVENTS REPORT")
    section_lines.append("=" * 60)

    if not excel_files:
        section_lines.append("Không tìm thấy file .xlsx nào.")
        return section_lines

    total_counter_domain = Counter()
    total_counter_ip = Counter()
    total_counter_url = Counter()
    total_ip_nation_map = defaultdict(Counter)
    grand_total_rows = 0

    section_lines.append(f"Tổng hợp {len(excel_files)} file .xlsx")
    section_lines.append("")

    for file_path in excel_files:
        try:
            counter_domain, counter_ip, counter_url, ip_nation_map, file_rows = process_excel_file(file_path)

            total_counter_domain.update(counter_domain)
            total_counter_ip.update(counter_ip)
            total_counter_url.update(counter_url)
            grand_total_rows += file_rows

            for ip, nation_counter in ip_nation_map.items():
                total_ip_nation_map[ip].update(nation_counter)

            section_lines.append(f"Đã xử lý: {file_path} | Số event: {file_rows}")
        except Exception as e:
            section_lines.append(f"Lỗi khi đọc file {file_path}: {e}")

    top_ip_data = get_top_ip_with_nation(
        total_counter_ip,
        total_ip_nation_map,
        top_k,
        grand_total_rows,
    )

    distinct_domain_count = len(total_counter_domain)

    section_lines.append("")
    section_lines.append(f"TỔNG số events ở tất cả file: {grand_total_rows}")
    section_lines.append(f"Số domain bị tấn công: {distinct_domain_count}")
    section_lines.append("")
    section_lines.extend(
        format_top(
            total_counter_domain,
            f"TOP {top_k} domain bị tấn công",
            top_k,
        )
    )
    section_lines.append("")
    section_lines.extend(format_top_ip_with_nation(top_ip_data))
    section_lines.append("")
    section_lines.extend(format_top(total_counter_url, f"TOP {top_k} URLs", top_k))

    return section_lines


def handle_attacks(folder_path: Path):
    pdf_files = list(folder_path.rglob("*.pdf"))

    section_lines = []
    section_lines.append("=" * 60)
    section_lines.append("ATTACKS REPORT")
    section_lines.append("=" * 60)

    if not pdf_files:
        section_lines.append("Không tìm thấy file .pdf nào.")
        return section_lines

    grand_total_web = 0
    grand_total_ddos = 0
    web_attack_counter = Counter()
    ddos_attack_counter = Counter()

    section_lines.append(f"Tổng hợp {len(pdf_files)} file .pdf")
    section_lines.append("")

    for file_path in pdf_files:
        try:
            result = process_pdf_file(file_path)

            grand_total_web += result["web_total"]
            grand_total_ddos += result["ddos_total"]

            for item in result["web_attack_types"]:
                web_attack_counter[item["type"]] += item["count"]

            for item in result["ddos_attack_types"]:
                ddos_attack_counter[item["type"]] += item["count"]

            section_lines.append(
                f"Đã xử lý: {file_path} | Web attacks: {result['web_total']} | DDoS attacks: {result['ddos_total']}"
            )
        except Exception as e:
            section_lines.append(f"Lỗi khi đọc file {file_path}: {e}")

    section_lines.append("")
    section_lines.append(f"TỔNG số cuộc tấn công lỗ hổng web: {grand_total_web}")
    section_lines.append(f"TỔNG số cuộc tấn công DDoS: {grand_total_ddos}")
    section_lines.append("")
    section_lines.extend(
        format_all_attack_types(
            "Các loại tấn công khai thác lỗ hổng",
            web_attack_counter,
        )
    )
    section_lines.append("")
    section_lines.extend(
        format_all_attack_types(
            "Các loại tấn công DDoS",
            ddos_attack_counter,
        )
    )

    return section_lines


def main():
    folder_path, top_k = parse_args()
    output_file = build_output_file(folder_path)

    report_lines = []
    report_lines.append(f"THƯ MỤC BÁO CÁO: {folder_path}")
    report_lines.append("")

    events_lines = handle_events(folder_path, top_k)
    attacks_lines = handle_attacks(folder_path)

    report_lines.extend(events_lines)
    report_lines.append("")
    report_lines.extend(attacks_lines)
    report_lines.append("")

    for line in report_lines:
        print(line)

    with open(output_file, "w", encoding="utf-8") as f:
        for line in report_lines:
            f.write(line + "\n")

    print(f"\nĐã ghi report ra file: {output_file}")


if __name__ == "__main__":
    main()